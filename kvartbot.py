# Для работы кода нужен Python3, а так же введите эти команды в консоль!
# pip install vk-api
# pip install random2
# pip install openpyxl
# pip install requests
# pip install glob2
# pip install os-sys
# pip install times
import vk_api # vk-api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api import VkUpload
import random # random2
import openpyxl # openpyxl
import requests # requests
import time 
from openpyxl import load_workbook
from glob import glob # glob2
import os # os-sys
import datetime # times
import traceback
while True:
	try:
		# Удаление всего из папки (пример: DOC/*.doc)
		def delite(name_files):
			files = glob(name_files)
			for f in files:
				os.remove(f)
		def msg(user_id, message, keys):
			vk.method('messages.send', {'peer_id': user_id,
										'message': str(message),
										'random_id': random.randint(1, 2912731987),
										'keyboard': keys})

		# Отправка файла
		def write_msg_upload(user_id, message, attachments):
			msg_upload(user_id, message, get_keys_delite(), attachments)
		def msg_upload(user_id, message, keys, attachments):
			vk.method('messages.send', {'peer_id': user_id,
										'message': str(message),
										'random_id': random.randint(1, 2912731987),
										'keyboard': keys,
										'attachment': ','.join(attachments)})

		def write_msg(user_id, message):
			msg(user_id, message, get_keys())

		def write_msg_no(user_id, message):
			msg(user_id, message, get_keys_no())

		def write_msg_object(user_id, message):
			msg(user_id, message, get_keys_object())

		def write_msg_day(user_id, message):
			msg(user_id, message, get_keys_day())

		def write_msg_pm(user_id, message):
			msg(user_id, message, get_keys_pm())

		def write_msg_pm_onli(user_id, message):
			msg(user_id, message, get_keys_pm_onli())

		def write_msg_naim(user_id, message):
			msg(user_id, message, get_keys_naim())

		def write_msg_yn(user_id, message):
			msg(user_id, message, get_keys_yn())

		def write_msg_ynu(user_id, message):
			msg(user_id, message, get_keys_ynu())

		def write_msg_access(user_id, message):
			msg(user_id, message, get_keys_access())

		def write_msg_universal(user_id, message, kol_vo):
			msg(user_id, message, get_keys_universal(kol_vo))

		def write_msg_redact(user_id, message):
			msg(user_id, message, get_keys_redact())

		def get_keys():
			menu = VkKeyboard()
			menu.add_button(label="Вернуться", color="negative")
			#menu.add_line()
			menu = menu.get_keyboard()
			return menu

		def get_keys_delite():
			menu = VkKeyboard(inline=True)
			menu.add_button(label="Стереть", color="negative")
			menu = menu.get_keyboard()
			return menu

		def get_keys_no():
			menu = VkKeyboard()
			menu.add_button(label="Моисеева", color="primary")
			menu.add_button(label="Переверткина", color="primary")
			menu.add_button(label="Ворошилова", color="primary")
			menu.add_line()
			menu.add_button(label="20 лет", color="primary")
			menu.add_button(label="Шишкова", color="primary")
			menu.add_button(label="Красноармейская", color="primary")
			menu.add_button(label="118", color="primary")
			menu.add_line()
			menu.add_button(label="За сегодня", color="positive")
			menu.add_button(label="За вчера", color="positive")
			menu.add_button(label="За позавчера", color="positive")
			menu.add_button(label="За 3", color="positive")
			menu.add_line()
			menu.add_button(label="Добавить расход на все", color="positive")
			# menu.add_button(label="Редактировать", color="negative")
			menu.add_button(label="Посмотреть", color="positive")
			menu = menu.get_keyboard()
			return menu
		def get_keys_object():
			menu = VkKeyboard()
			menu.add_button(label="Моисеева", color="primary")
			menu.add_button(label="Переверткина", color="primary")
			menu.add_button(label="Ворошилова", color="primary")
			menu.add_line()
			menu.add_button(label="20 лет", color="primary")
			menu.add_button(label="Шишкова", color="primary")
			menu.add_button(label="Красноармейская", color="primary")
			menu.add_button(label="118", color="primary")
			menu.add_line()
			menu.add_button(label="Вернуться", color="negative")
			menu = menu.get_keyboard()
			return menu
		def get_keys_pm():
			menu = VkKeyboard()
			menu.add_button(label="+", color="primary")
			menu.add_button(label="-", color="primary")
			menu.add_line()
			menu.add_button(label="Код", color="positive")
			menu.add_line()
			menu.add_button(label="Посмотреть", color="positive")
			menu.add_line()
			menu.add_button(label="Вернуться", color="negative")
			menu = menu.get_keyboard()
			return menu
		def get_keys_pm_onli():
			menu = VkKeyboard()
			menu.add_button(label="+", color="primary")
			menu.add_button(label="-", color="primary")
			menu.add_line()
			menu.add_button(label="Вернуться", color="negative")
			menu = menu.get_keyboard()
			return menu
		def get_keys_day():
			menu = VkKeyboard()
			menu.add_button(label="Сегодня", color="primary")
			menu.add_button(label="Вчера", color="primary")
			menu.add_button(label="Позавчера", color="primary")
			menu.add_line()
			menu.add_button(label="Вернуться", color="negative")
			menu = menu.get_keyboard()
			return menu
		def get_keys_yn():
			menu = VkKeyboard()
			menu.add_button(label="Да", color="positive")
			menu.add_button(label="Нет", color="negative")
			menu = menu.get_keyboard()
			return menu
		def get_keys_ynu():
			menu = VkKeyboard()
			menu.add_button(label="Да", color="positive")
			menu.add_button(label="Нет", color="negative")
			menu.add_line()
			menu.add_button(label="Указать свой", color="primary")
			menu = menu.get_keyboard()
			return menu
		def get_keys_naim():
			menu = VkKeyboard(inline=True)
			menu.add_button(label="Авито", color="primary")
			menu.add_button(label="Коммуналка", color="primary")
			menu.add_button(label="Транспорт", color="primary")
			menu.add_line()
			menu.add_button(label="Прочее", color="primary")
			menu.add_button(label="Уборка", color="primary")
			menu.add_button(label="Квартплата", color="primary")
			menu = menu.get_keyboard()
			return menu
		def get_keys_access():
			menu = VkKeyboard()
			menu.add_button(label="Укажите код", color="negative")
			menu = menu.get_keyboard()
			return menu
		def get_keys_universal(kol_vo):
			menu = VkKeyboard()
			i = 1
			for g in range(kol_vo):
				if i == 7:
					i = 1
					menu.add_line()
				menu.add_button(label=str(int(g)+1), color="primary")
				i += 1

			menu.add_line()
			menu.add_button(label="Вернуться", color="negative")
			menu = menu.get_keyboard()
			return menu
		def get_keys_redact():
			menu = VkKeyboard()
			menu.add_button(label="Редактировать", color="positive")
			menu.add_button(label="Удалить", color="negative")
			menu.add_line()
			menu.add_button(label="Вернуться", color="negative")
			menu = menu.get_keyboard()
			return menu
		#Узнаём что человек хочет сделать
		def poisk_etap(ws, name, access):
			n = 1
			while True:
				if ws['A'+str(n)].value == name:
					return (str(ws['B'+str(n)].value), n)
				elif ws['A'+str(n)].value == None:
					ws['A'+str(n)].value = name
					if access == True:
						ws['B'+str(n)].value = 'NO'
					else:
						ws['B'+str(n)].value = 'NOACCESS'
					wb.save('kvartbot/acc.xlsx')
					n -= 1
				n += 1

		def etap_NO(name, n, ws):
			write_msg_no(name, 'Тыкай по кнопкам')
			ws['B'+str(n)].value = '1'
			#delite('mukulatur/' + str(event.user_id) + '.txt')

		def cheng_etap(ws, etap, znach):
			ws['B'+str(nomer)].value = etap + '.' + str(znach)

		def searth_report(what_day):
			global moscow
			if what_day == 1:
				day = datetime.datetime.now(moscow).strftime('%d.%m.%Y')
			elif what_day == 2:
				day = (datetime.datetime.now(moscow)-datetime.timedelta(days=1)).strftime('%d.%m.%Y')
			elif what_day == 3:
				day = (datetime.datetime.now(moscow)-datetime.timedelta(days=2)).strftime('%d.%m.%Y')
			elif what_day == 4:
				day = (datetime.datetime.now(moscow)-datetime.timedelta(days=3)).strftime('%d.%m.%Y')
			report = ''
			report = searth_in_wb(report, 'moi.xlsx', day)
			report = searth_in_wb(report, 'per.xlsx', day)
			report = searth_in_wb(report, 'vor.xlsx', day)
			#report = searth_in_wb(report, 'kor.xlsx', day)
			report = searth_in_wb(report, 'let.xlsx', day)
			report = searth_in_wb(report, 'shish.xlsx', day)
			report = searth_in_wb(report, 'kra.xlsx', day)
			report = searth_in_wb(report, '118.xlsx', day)
			return report

		def searth_in_wb(report, name_wb, day):
			def get_plus(ws, plus, day):
				i = 1
				while ws['A'+str(i)].value != None:
					if str(ws['A'+str(i)].value) == str(day):
						plus = plus + str(ws['G'+str(i)].value) + ' &#10134; ' + str(ws['B'+str(i)].value) + ' * ' + str(ws['C'+str(i)].value) + ' = ' + str(int(ws['B'+str(i)].value) * int(ws['C'+str(i)].value)) + '\n'
					i += 1
				return plus

			def get_minus(ws, minus, day):
				i = 1
				while ws['D'+str(i)].value != None:
					if ws['D'+str(i)].value == str(day):
						minus = minus + str(ws['H'+str(i)].value) + ' &#10134; ' + str(ws['E'+str(i)].value) + ' = ' + str(ws['F'+str(i)].value) + '\n'
					i += 1
				return minus
				
			wb = load_workbook(filename = 'kvartbot/'+name_wb)
			ws = wb.active
			if name_wb == 'moi.xlsx':
				name_wb = 'Моисеева'
			elif name_wb == 'per.xlsx':
				name_wb = 'Переверткина'
			elif name_wb == 'vor.xlsx':
				name_wb = 'Ворошилова'
			elif name_wb == 'kor.xlsx':
				name_wb = 'Короленко'
			elif name_wb == 'let.xlsx':
				name_wb = '20 лет'
			elif name_wb == 'shish.xlsx':
				name_wb = 'Шишкова'
			elif name_wb == 'kra.xlsx':
				name_wb = 'Красноармейская'
			elif name_wb == '118.xlsx':
				name_wb = '118'

			plus = ''
			minus = ''
			plus = get_plus(ws, plus, day)
			minus = get_minus(ws, minus, day)
			if plus != '' or minus != '':
				report = report + '\n&#9193;' + name_wb + '&#9194;\n'
				if plus != '':
					report = report + 'Доходы:\n' + plus
				if minus != '':
					report = report + 'Расходы:\n' + minus
			wb.close()
			return report

		# цена за сутки
		def money(ws, etap, event, event_id):
			global day
			cheng_etap(ws, etap, 1)
			file = open(str(event_id) + ".txt", "w")
			file.write(str(day) + '\n')
			file.write(str(int(event)) + '\n')
			file.close()

		# кол-во суток и согранение
		def plus(ws, etap, xl, event, event_id):
			global time_now
			if xl == 1:
				filename = 'kvartbot/moi.xlsx'
			elif xl == 2:
				filename = 'kvartbot/per.xlsx'
			elif xl == 3:
				filename = 'kvartbot/vor.xlsx'
			elif xl == 4:
				filename = 'kvartbot/kor.xlsx'
			elif xl == 5:
				filename = 'kvartbot/let.xlsx'
			elif xl == 6:
				filename = 'kvartbot/shish.xlsx'
			elif xl == 7:
				filename = 'kvartbot/kra.xlsx'
			elif xl == 8:
				filename = 'kvartbot/118.xlsx'

			file = open(str(event_id) + '.txt', 'a')
			file.write(str(int(event)) + '\n')
			file.close()
			wb1 = load_workbook(filename = filename)
			ws1 = wb1.active
			i = 1
			while True:
				if ws1['A'+str(i)].value != None:
					i += 1
				else:
					
					file = open(str(event_id) + '.txt', 'r')
					line = []
					for line_all in file:
						line.append(line_all)
					ws1['A'+str(i)].value = str(line[0].replace('\n', ''))
					ws1['B'+str(i)].value = str(line[1].replace('\n', ''))
					ws1['C'+str(i)].value = str(line[2].replace('\n', ''))
					ws1['G'+str(i)].value = str(time_now)
					file.close()
					wb1.save(filename)
					break

			write_msg_pm(event_id, 'Сохраненно')
			delite(str(event_id) + '.txt')

		
		# наименование и сохранение
		def minus(ws, etap, xl, event, event_id):
			global time_now
			if xl == 1:
				filename = 'kvartbot/moi.xlsx'
			elif xl == 2:
				filename = 'kvartbot/per.xlsx'
			elif xl == 3:
				filename = 'kvartbot/vor.xlsx'
			elif xl == 4:
				filename = 'kvartbot/kor.xlsx'
			elif xl == 5:
				filename = 'kvartbot/let.xlsx'
			elif xl == 6:
				filename = 'kvartbot/shish.xlsx'
			elif xl == 7:
				filename = 'kvartbot/kra.xlsx'
			elif xl == 8:
				filename = 'kvartbot/118.xlsx'

			file = open(str(event_id) + '.txt', 'a')
			file.write(str(event) + '\n')
			file.close()
			wb1 = load_workbook(filename = filename)
			ws1 = wb1.active
			i = 1		
			file = open(str(event_id) + '.txt', 'r')
			line = []
			for line_all in file:
				line.append(line_all)
			while ws1['D'+str(i)].value != None:
				i += 1
			# line[0] - Дата(D)	line[1] - цена(F) line[2] - наим(E)
			ws1['D'+str(i)].value = str(line[0].replace('\n', ''))
			ws1['F'+str(i)].value = str(line[1].replace('\n', ''))
			ws1['E'+str(i)].value = str(line[2].replace('\n', ''))
			ws1['H'+str(i)].value = str(time_now)
			file.close()
			wb1.save(filename)
			write_msg_pm(event_id, 'Сохраненно')
			delite(str(event_id) + '.txt')

		def minus_all(event):
			global time_now
			global amount_kvart
			for xl in range(amount_kvart):  
				xl += 1
				if xl == 1:
					filename = 'kvartbot/moi.xlsx'
				elif xl == 2:
					filename = 'kvartbot/per.xlsx'
				elif xl == 3:
					filename = 'kvartbot/vor.xlsx'
				elif xl == 4:
					filename = 'kvartbot/let.xlsx'
				elif xl == 5:
					filename = 'kvartbot/shish.xlsx'
				elif xl == 6:
					filename = 'kvartbot/kra.xlsx'
				elif xl == 7:
					filename = 'kvartbot/118.xlsx'
				file = open(str(event.user_id) + '.txt', 'a')
				file.write(str(event.text) + '\n')
				file.close()
				wb1 = load_workbook(filename = filename)
				ws1 = wb1.active
				i = 1		
				file = open(str(event.user_id) + '.txt', 'r')
				line = []
				for line_all in file:
					line.append(line_all)
				while ws1['D'+str(i)].value != None:
					i += 1
				ws1['D'+str(i)].value = str(line[0].replace('\n', ''))
				ws1['F'+str(i)].value = int(int(line[1].replace('\n', ''))/amount_kvart+1)
				ws1['E'+str(i)].value = str(line[2].replace('\n', ''))
				ws1['H'+str(i)].value = str(time_now)	
					
				file.close()
				wb1.save(filename)
			write_msg(event.user_id, 'Сохраненно.')

		def proverka_xl():
			global amount_kvart
			for i in range(amount_kvart):
				i += 1
				if i == 1:
					filename = 'kvartbot/moi.xlsx'
				elif i == 2:
					filename = 'kvartbot/per.xlsx'
				elif i == 3:
					filename = 'kvartbot/vor.xlsx'
				elif i == 4:
					filename = 'kvartbot/kra.xlsx'
				elif i == 5:
					filename = 'kvartbot/let.xlsx'
				elif i == 6:
					filename = 'kvartbot/shish.xlsx'
				elif i == 7:
					filename = 'kvartbot/118.xlsx'

				if os.path.exists(filename) == False:
					wb1 = openpyxl.Workbook()
					ws1 = wb1.active
					ws1['A1'].value = 'Дата'
					ws1['B1'].value = 'За сутки'
					ws1['C1'].value = 'Кол-во суток'
					ws1['D1'].value = 'Дата'
					ws1['E1'].value = 'Наим'
					ws1['F1'].value = 'Цена'
					ws1['G1'].value = 'Время в +'
					ws1['H1'].value = 'Время в -'

					# temp = 'SUMPRODUCT(B:B;C:C)'
					# ws1['G2'].value = '=' + temp
					# temp = 'SUM(F:F)'
					# ws1['H2'].value = '=' + temp
					# temp = 'G2-H2'
					# ws1['I2'].value = '=' + temp	
					wb1.save(filename)
					wb1.close()

		def rezult(wbg, wsg, name_xl):
			global income_all
			global day_all
			wb1 = openpyxl.Workbook()
			ws1 = wb1.active
			ws1['A1'].value = 'Дата'
			ws1['B1'].value = 'За сутки'
			ws1['C1'].value = 'Кол-во суток'
			ws1['D1'].value = 'Дата'
			ws1['E1'].value = 'Наим'
			ws1['F1'].value = 'Цена'
			ws1['G1'].value = 'Прибыль'
			ws1['H1'].value = 'Затраты'
			ws1['I1'].value = 'Доход'
			ws1['J1'].value = 'Кол-во занятых суток'



			# переписываем все записи с доходами 
			i = 2	
			while wsg['A' + str(i)].value != None:
				ws1['A' + str(i)].value = wsg['A' + str(i)].value
				ws1['B' + str(i)].value = int(wsg['B' + str(i)].value)
				ws1['C' + str(i)].value = int(wsg['C' + str(i)].value)
				i += 1


			# Считаем прибыль
			i = 2 
			profit = 0
			while ws1['B' + str(i)].value != None:
				profit = profit + int(wsg['B' + str(i)].value) * int(wsg['C' + str(i)].value)
				i += 1
			ws1['G' + str(2)].value = profit

			#считаем общее кол-во суток
			i = 2
			g = 0
			while ws1['C' + str(i)].value != None:
				g = g + int(ws1['C' + str(i)].value) 
				i += 1  
			day_all = day_all + g
			ws1['J' + str(2)].value = g

			# Переписываем расходы
			# ставим курсор в начало в старом и в отпрвляемом
			i = 2
			g = 2
			expenses = 0
			# Создание пустого списка под названия которые были	
			lis = []
			# Пока в старом есть записи
			while wsg['D' + str(g)].value != None:
				# Записываем название из старого наименования затрат
				temp = str(wsg['E' + str(g)].value)
				# Проверяем название было ли оно раньше
				if temp not in lis:
					summ = 0
					c = g
					# Сохраняем название в список
					lis.append(temp)
					# Проверяем все названия в старом пока они там есть 
					while wsg['D' + str(c)].value != None:
						if temp == wsg['E' + str(c)].value:
							summ = summ + int(wsg['F' + str(c)].value)
							date_last = wsg['D' + str(c)].value
						c += 1
					ws1['D' + str(i)].value = str(date_last)
					ws1['E' + str(i)].value = temp
					ws1['F' + str(i)].value = summ
					expenses = expenses + summ
					i += 1
				g += 1
			
			income = profit - expenses
			income_all = income_all + income
			ws1['H' + str(2)].value = expenses
			ws1['I' + str(2)].value = income
			
			wb1.save(name_xl.split('/')[0] + '/temp_' + name_xl.split('/')[1])
			wb1.close()
			return (name_xl.split('/')[0] + '/temp_' + name_xl.split('/')[1])
			# i = 2 
			# profit = 0
			# while wsg['B' + str(i)].value != None:
			# 	profit = profit + int(wsg['B' + str(i)].value) * int(wsg['C' + str(i)].value)
			# 	i += 1
			# wsg['G' + str(2)].value = profit
			# i = 2
			# expenses = 0
			# while wsg['F' + str(i)].value != None:
			# 	expenses = expenses + cell_amount(wsg['F' + str(i)].value)
			# 	i += 1
			# wsg['H' + str(2)].value = expenses
			# income = profit - expenses
			# wsg['I' + str(2)].value = income
			# wbg.save(name_xl)

		def uppload_xlsx(fail, attachments, zapisi):
			# убираем kvartbot
			name_fail = fail.split('/')[1]
			# Открываем фаил
			wbg = load_workbook(filename = fail)
			wsg = wbg.active
			# Если он не пустой
			if wsg['A2'].value != None or wsg['D2'].value != None:
				name_uppload = rezult(wbg, wsg, fail)
				upload_file = upload.document_message(doc = name_uppload, peer_id = event.user_id, title = name_fail)['doc']
				attachments.append('doc{}_{}'.format(upload_file['owner_id'], upload_file['id']))
				zapisi = True
				delite(name_uppload)
			wbg.close()

			return attachments, zapisi
		

		def cell_amount(cell_old):
			cell = str(cell_old).replace('=', '')
			cell = cell.replace(' ', '')
			cell = cell.replace('+', ' ')
			rez = 0
			for elem in cell.split():
				rez = rez + int(elem)
			return(rez)

		def transformation(income_all):
			
			rezult = '{0:,}'.format(income_all).replace(',', ' ')

			return rezult

		def exam(number):
			try:
				if int(number) > 0 and int(number) < 20:
					return True
				else:
					return False
			except Exception as e:
				return False

		def edit(name_wb, day, pm):
			global moscow
			if name_wb == 'Моисеева':
				name_wb = 'moi.xlsx'
			elif name_wb == 'Переверткина':
				name_wb = 'per.xlsx'
			elif name_wb == 'Ворошилова':
				name_wb = 'vor.xlsx'
			elif name_wb == 'Короленко':
				name_wb = 'kor.xlsx'
			elif name_wb == '20 лет':
				name_wb = 'let.xlsx'
			elif name_wb == 'Шишкова':
				name_wb = 'shish.xlsx'
			elif name_wb == 'Красноармейская':
				name_wb = 'kra.xlsx'
			elif name_wb == '118':
				name_wb = '118.xlsx'

			if day == 'Сегодня':
				day = datetime.datetime.now(moscow).strftime('%d.%m.%Y')
			elif day == 'Вчера':
				day = (datetime.datetime.now(moscow)-datetime.timedelta(days=1)).strftime('%d.%m.%Y')
			elif day == 'Позавчера':
				day = (datetime.datetime.now(moscow)-datetime.timedelta(days=2)).strftime('%d.%m.%Y')

			wb = load_workbook(filename = 'kvartbot/'+name_wb)
			ws = wb.active
			if pm == '+':
				rezult, kol_vo = get_plus_edit(ws, day)
			elif pm  == '-':
				rezult, kol_vo = get_minus_edit(ws, day)
			wb.close()
			if rezult == '':
				return 'Нет записей', kol_vo
			else:
				return 'Выберите цифрой что редактировать \n' + rezult, kol_vo

		def get_plus_edit(ws, day):
			i = 1
			g = 0
			rezult = ''
			while ws['A'+str(i)].value != None:
				if str(ws['A'+str(i)].value) == str(day):
					g += 1
					rezult = rezult + str(g) + ') ' + str(ws['A'+str(i)].value) + ' ' + str(ws['B'+str(i)].value) + ' ' + str(ws['C'+str(i)].value) + '\n'
				i += 1
			return rezult, g

		def get_minus_edit(ws, day):
			i = 1
			g = 0
			rezult = ''
			while ws['D'+str(i)].value != None:
				if str(ws['D'+str(i)].value) == str(day):
					g += 1
					rezult = rezult + str(g) + ') ' + str(ws['D'+str(i)].value) + ' ' + str(ws['E'+str(i)].value) + ' ' + str(ws['F'+str(i)].value) + '\n'
				i += 1
			return rezult, g

		token = "a95404fb4302cef1f1b215034bc1dfd270072d044b4658c823040d31677fd0b0b7e3e62bc5a9e68eb9c24"
		# Авторизуемся как сообщество
		vk = vk_api.VkApi(token=token)
		# Работа с сообщениями
		longpoll = VkLongPoll(vk)
		upload = VkUpload(vk)
		#Кол-во квартир
		amount_kvart = 7
		# Создание тхт с тем кто может пользоваться ботом
		admin_id = ['156873624', '145665746']
		if os.path.exists('kvartbot/admin_id.txt') == False:
			f = open('kvartbot/admin_id.txt', 'w')
			for line in admin_id:
				f.write(line + '\n')
			f.close()

		if os.path.exists('kvartbot/code_moi.txt') == False:
			f = open('kvartbot/code_moi.txt', 'w')
			f.write('отсутствует')
			f.close()
		f = open('kvartbot/code_moi.txt', 'r')
		for line in f:
			code_moi = str(line)
		f.close()	
		if os.path.exists('kvartbot/code_per.txt') == False:
			f = open('kvartbot/code_per.txt', 'w')
			f.write('отсутствует')
			f.close()
		f = open('kvartbot/code_per.txt', 'r')
		for line in f:
			code_per = str(line)
		f.close()
		if os.path.exists('kvartbot/code_vor.txt') == False:
			f = open('kvartbot/code_vor.txt', 'w')
			f.write('отсутствует')
			f.close()
		f = open('kvartbot/code_vor.txt', 'r')
		for line in f:
			code_vor = str(line)
		f.close()
		# if os.path.exists('kvartbot/code_kor.txt') == False:
		# 	f = open('kvartbot/code_kor.txt', 'w')
		# 	f.write('отсутствует')
		# 	f.close()
		# f = open('kvartbot/code_kor.txt', 'r')
		# for line in f:
		# 	code_kor = str(line)
		# f.close()
		if os.path.exists('kvartbot/code_let.txt') == False:
			f = open('kvartbot/code_let.txt', 'w')
			f.write('отсутствует')
			f.close()
		f = open('kvartbot/code_let.txt', 'r')
		for line in f:
			code_let = str(line)
		f.close()
		if os.path.exists('kvartbot/code_shish.txt') == False:
			f = open('kvartbot/code_shish.txt', 'w')
			f.write('отсутствует')
			f.close()
		f = open('kvartbot/code_shish.txt', 'r')
		for line in f:
			code_shish = str(line)
		f.close()
		if os.path.exists('kvartbot/code_kra.txt') == False:
			f = open('kvartbot/code_kra.txt', 'w')
			f.write('отсутствует')
			f.close()
		f = open('kvartbot/code_kra.txt', 'r')
		for line in f:
			code_kra = str(line)
		f.close()
		if os.path.exists('kvartbot/code_118.txt') == False:
			f = open('kvartbot/code_118.txt', 'w')
			f.write('отсутствует')
			f.close()
		f = open('kvartbot/code_118.txt', 'r')
		for line in f:
			code_118 = str(line)
		f.close()
		print('Запущен!')
		for event in longpoll.listen():
			# Если пришло новое сообщение
			if event.type == VkEventType.MESSAGE_NEW:
				if event.to_me:
					access = False
					f = open('kvartbot/admin_id.txt', 'r')
					for line in f:
						if str(event.user_id) + '\n' == str(line):
							access = True
							
					f.close()	
					
					# Определение какой сегодня день
					moscow = datetime.timezone(datetime.timedelta(hours = 3))
					day = datetime.datetime.now(moscow).strftime('%d.%m.%Y')
					time_now = datetime.datetime.now(moscow).strftime('%H:%M')
					# Проверка наличия файлов
					if os.path.exists('kvartbot/acc.xlsx') == False:
						wb = openpyxl.Workbook()			
						wb.save('kvartbot/acc.xlsx')
						wb.close()

					proverka_xl()
					income_all = 0
					day_all = 0
					#Поиск того, на каком этапе находится пользователь
					wb = load_workbook(filename = 'kvartbot/acc.xlsx')
					ws = wb.active
					#date = ws['C'+str(1)].value
					etap, nomer = poisk_etap(ws, event.user_id, access)
					zapisi = True
					zapisi1 = True
					if etap == 'NO':
						etap_NO(event.user_id, nomer, ws)
					#--------------------------------------------------------------------------------------------------------------------------
					# основное меню
					elif etap == '1':
						if event.text == 'Моисеева':
							ws['B'+str(nomer)].value = '1.1'
							write_msg_pm(event.user_id, 'Прибыль или затраты?')

						elif event.text == 'Переверткина':
							ws['B'+str(nomer)].value = '1.2'
							write_msg_pm(event.user_id, 'Прибыль или затраты?')

						elif event.text == 'Ворошилова':
							ws['B'+str(nomer)].value = '1.3'
							write_msg_pm(event.user_id, 'Прибыль или затраты?')

						elif event.text == 'Короленко':
							write_msg_no(event.user_id, 'Её больше нет')

						elif event.text == '20 лет':
							ws['B'+str(nomer)].value = '1.5'
							write_msg_pm(event.user_id, 'Прибыль или затраты?')

						elif event.text == 'Шишкова':
							ws['B'+str(nomer)].value = '1.6'
							write_msg_pm(event.user_id, 'Прибыль или затраты?')

						elif event.text == 'Красноармейская':
							ws['B'+str(nomer)].value = '1.7'
							write_msg_pm(event.user_id, 'Прибыль или затраты?')

						elif event.text == '118':
							ws['B'+str(nomer)].value = '1.8'
							write_msg_pm(event.user_id, 'Прибыль или затраты?')

						elif event.text == 'За сегодня': 
							report = searth_report(1)
							if report == '':
								write_msg_no(event.user_id, 'Записей за сегодня нет')
							else:
								write_msg_no(event.user_id, 'Записи за сегодня: ' + '\n' + report)

						elif event.text == 'За вчера': 
							report = searth_report(2)
							if report == '':
								write_msg_no(event.user_id, 'Записей за вчера нет')
							else:
								write_msg_no(event.user_id, 'Записи за вчера: ' + '\n' + report)

						elif event.text == 'За позавчера': 
							report = searth_report(3)
							if report == '':
								write_msg_no(event.user_id, 'Записей за позавчера нет')
							else:
								write_msg_no(event.user_id, 'Записи за позавчера: ' + '\n' + report)

						elif event.text == 'За 3': 
							report = searth_report(4)
							if report == '':
								write_msg_no(event.user_id, 'Записей за 3 дня нет')
							else:
								write_msg_no(event.user_id, 'Записи за 3 дня: ' + '\n' + report)

						elif event.text == 'Добавить расход на все':
							ws['B'+str(nomer)].value = '0.2'
							write_msg(event.user_id, 'Укажи цену')

						elif event.text == 'Редактировать':
							write_msg_object(event.user_id, 'Выберите объект')
							ws['B'+str(nomer)].value = 'choice_edit'

						elif event.text == 'Посмотреть':
							attachments = []
							zapisi = False
							attachments, zapisi = uppload_xlsx('kvartbot/moi.xlsx', attachments, zapisi)
							attachments, zapisi = uppload_xlsx('kvartbot/per.xlsx', attachments, zapisi)
							attachments, zapisi = uppload_xlsx('kvartbot/vor.xlsx', attachments, zapisi)
							# attachments, zapisi = uppload_xlsx('kvartbot/kor.xlsx', attachments, zapisi)
							attachments, zapisi = uppload_xlsx('kvartbot/let.xlsx', attachments, zapisi)
							attachments, zapisi = uppload_xlsx('kvartbot/shish.xlsx', attachments, zapisi)
							attachments, zapisi = uppload_xlsx('kvartbot/kra.xlsx', attachments, zapisi)
							attachments, zapisi = uppload_xlsx('kvartbot/118.xlsx', attachments, zapisi)
							if zapisi == True:
								write_msg_upload(event.user_id, 'Общий доход: ' + str(transformation(income_all)) + '\nНочей сдано: ' + str(day_all), attachments)
							else:
								write_msg_no(event.user_id, 'Записей нет')

							
						elif event.text == 'Стереть':
							write_msg_yn(event.user_id, 'Уверен?')
							ws['B'+str(nomer)].value = 'del'

						elif event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)

						else:
							write_msg_no(event.user_id, 'Тыкай по кнопкам!')
					#--------------------------------------------------------------------------------------------------------------------------
					elif etap == 'del':
						if event.text == 'Да':
							delite('kvartbot/moi.xlsx')
							delite('kvartbot/per.xlsx')
							delite('kvartbot/vor.xlsx')
							delite('kvartbot/let.xlsx')
							delite('kvartbot/shish.xlsx')
							delite('kvartbot/kra.xlsx')
							delite('kvartbot/118.xlsx')
							etap_NO(event.user_id, nomer, ws)
						elif event.text == 'Нет':
							etap_NO(event.user_id, nomer, ws)
						else:
							write_msg_yn(event.user_id, 'Я тебя не понял')
					#--------------------------------------------------------------------------------------------------------------------------
					# после выбора квартиры
					elif etap == '1.1' or etap == '1.2' or etap == '1.3' or etap == '1.4' or etap == '1.5' or etap == '1.6' or etap == '1.7' or etap == '1.8':
						if etap == '1.1':
							fail = 'kvartbot/moi.xlsx'
						elif etap == '1.2':
							fail = 'kvartbot/per.xlsx'
						elif etap == '1.3':
							fail = 'kvartbot/vor.xlsx'
						elif etap == '1.4':
							fail = 'kvartbot/kor.xlsx'
						elif etap == '1.5':
							fail = 'kvartbot/let.xlsx'
						elif etap == '1.6':
							fail = 'kvartbot/shish.xlsx'
						elif etap == '1.7':
							fail = 'kvartbot/kra.xlsx'
						elif etap == '1.8':
							fail = 'kvartbot/118.xlsx'
						if event.text == '+':
							write_msg (event.user_id, 'Укажи цену за сутки')
							cheng_etap(ws, etap, 1)
						elif event.text == '-':
							write_msg (event.user_id, 'Укажи цену')
							cheng_etap(ws, etap, 2)
						elif event.text == 'Код':
							if etap == '1.1':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_moi.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. Моисеева, д. 69, 11 этаж, кв. 254\n' + 'Домофон: 254В\n' + 'Сейф: ' + str(code_moi))
								write_msg_ynu(event.user_id, 'Код от Моисеева.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod1'
							elif etap == '1.2':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_per.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. Переверткина, д. 31а, 3 под., 3 этаж, 91 кв.\n' + 'Домофон: В91В3021\n' + 'Сейф: ' + str(code_per))
								write_msg_ynu(event.user_id, 'Код от Переверткина 91.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod2'	
							elif etap == '1.3':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_vor.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. Ворошилова, д. 45б, 1 под., 8 этаж, 51 кв.\n' + 'Домофон: 51В\n' + 'Сейф: ' + str(code_vor))
								write_msg_ynu(event.user_id, 'Код от Ворошилова.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod3'
							elif etap == '1.4':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_kor.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. Короленко, д. 5, под. 2, этаж 4, 38 кв.\n' + 'Домофон: 38В\n' + 'Сейф: ' + str(code_kor))
								write_msg_ynu(event.user_id, 'Код от Короленко.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod4'
							elif etap == '1.5':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_let.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. 20 - летия Октября, д. 107, под. 3, этаж 1, кв. 42\n' + 'Домофон: 42\n' + 'Сейф: ' + str(code_let))
								write_msg_ynu(event.user_id, 'Код от 20 лет.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod5'
							elif etap == '1.6':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_shish.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. Шишкова, д. 144б, под. 3, этаж 7, кв. 201\n' + 'Домофон: 9753\n' + 'Сейф: ' + str(code_shish))
								write_msg_ynu(event.user_id, 'Код от Шишкова.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod6'
							elif etap == '1.7':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_kra.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. Красноармейская, д. 17, под. 7, этаж 1, кв. 103\n' + 'Домофон: 103\n' + 'Сейф: ' + str(code_kra))
								write_msg_ynu(event.user_id, 'Код от Красноармейской.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod7'
							elif etap == '1.8':
								last_chenge = datetime.datetime.fromtimestamp(os.path.getmtime('kvartbot/code_118.txt')).strftime('%d.%m.%Y')
								write_msg_ynu(event.user_id, 'Адрес: ул. Переверткина, д. 31а, под. 3, этаж 10, 118 кв.\n' + 'Домофон: В91В3021\n' + 'Сейф: ' + str(code_118))
								write_msg_ynu(event.user_id, 'Код от Переверткина 118.' + '\nПоследнее изменение: ' + str(last_chenge) + '.\nПоменять на рандомный?')
								ws['B'+str(nomer)].value = 'kod8'

						elif event.text == 'Посмотреть':
							attachments = []
							zapisi = False
							attachments, zapisi = uppload_xlsx(fail, attachments, zapisi)
							
							if zapisi == True:
								write_msg_upload(event.user_id, 'Доход: ' + str(transformation(income_all)) + '\nНочей сдано: ' + str(day_all), attachments)
							else:
								write_msg_pm(event.user_id, 'Записей нет')

						elif event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)

						elif event.text == 'Стереть':
							write_msg_yn(event.user_id, 'Уверен?')

						elif event.text == 'Да':
							delite(fail)
							write_msg_pm(event.user_id, '"' + str(fail.split('/')[1]) + '" - Удалён')
						else:
							write_msg_pm(event.user_id, 'Тыкай по кнопкам')

					elif etap == 'kod1' or etap == 'kod2' or etap == 'kod3' or etap == 'kod4' or etap == 'kod5' or etap == 'kod6' or etap == 'kod7' or etap == 'kod8':
						if event.text == 'Да':
							random_code = 000
							while int(len(str(random_code))) < 4: 
								random_code = random.randint(0000, 9999) 
							if etap == 'kod1':
								code_moi = random_code
								f = open('kvartbot/code_moi.txt', 'w')
								f.write(str(code_moi))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_moi))
							elif etap == 'kod2':
								code_per = random_code
								f = open('kvartbot/code_per.txt', 'w')
								f.write(str(code_per))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_per))
							elif etap == 'kod3':
								code_vor = random_code
								f = open('kvartbot/code_vor.txt', 'w')
								f.write(str(code_vor))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_vor))
							elif etap == 'kod4':
								code_kor = random_code
								f = open('kvartbot/code_kor.txt', 'w')
								f.write(str(code_kor))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_kor))
							elif etap == 'kod5':
								code_let = random_code
								f = open('kvartbot/code_let.txt', 'w')
								f.write(str(code_let))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_let))
							elif etap == 'kod6':
								code_shish = random_code
								f = open('kvartbot/code_shish.txt', 'w')
								f.write(str(code_shish))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_shish))
							elif etap == 'kod7':
								code_kra = random_code
								f = open('kvartbot/code_kra.txt', 'w')
								f.write(str(code_kra))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_kra))
							elif etap == 'kod8':
								code_118 = random_code
								f = open('kvartbot/code_118.txt', 'w')
								f.write(str(code_118))
								f.close()
								write_msg_no(event.user_id, 'Сохраненно, новый код: ' + str(code_118))
							ws['B'+str(nomer)].value = '1'

						elif event.text == 'Нет':
							etap_NO(event.user_id, nomer, ws)

						elif event.text == 'Указать свой':
							if etap == 'kod1':
								ws['B'+str(nomer)].value = 'kod1.1'
							elif etap == 'kod2':
								ws['B'+str(nomer)].value = 'kod1.2'
							elif etap == 'kod3':
								ws['B'+str(nomer)].value = 'kod1.3'
							elif etap == 'kod4':
								ws['B'+str(nomer)].value = 'kod1.4'
							elif etap == 'kod5':
								ws['B'+str(nomer)].value = 'kod1.5'
							elif etap == 'kod6':
								ws['B'+str(nomer)].value = 'kod1.6'
							elif etap == 'kod7':
								ws['B'+str(nomer)].value = 'kod1.7'
							elif etap == 'kod8':
								ws['B'+str(nomer)].value = 'kod1.8'
							write_msg(event.user_id, 'Укажите')
						else:
							write_msg_ynu(event.user_id, 'Я тебя не понял')

					elif etap == 'kod1.1' or etap == 'kod1.2' or etap == 'kod1.3' or etap == 'kod1.4' or etap == 'kod1.5' or etap == 'kod1.6' or etap == 'kod1.7' or etap == 'kod1.8':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						elif len(str(event.text)) == 4 and event.text.isdigit() == True:
							if etap == 'kod1.1':
								code_moi = event.text
								f = open('kvartbot/code_moi.txt', 'w')
								f.write(str(event.text))
								f.close()
							elif etap == 'kod1.2':
								code_per = event.text
								f = open('kvartbot/code_per.txt', 'w')
								f.write(str(code_per))
								f.close()
							elif etap == 'kod1.3':
								code_vor = event.text
								f = open('kvartbot/code_vor.txt', 'w')
								f.write(str(code_vor))
								f.close()
							elif etap == 'kod1.4':
								code_kor = event.text
								f = open('kvartbot/code_kor.txt', 'w')
								f.write(str(code_kor))
								f.close()
							elif etap == 'kod1.5':
								code_let = event.text
								f = open('kvartbot/code_let.txt', 'w')
								f.write(str(code_let))
								f.close()
							elif etap == 'kod1.6':
								code_shish = event.text
								f = open('kvartbot/code_shish.txt', 'w')
								f.write(str(code_shish))
								f.close()
							elif etap == 'kod1.7':
								code_kra = event.text
								f = open('kvartbot/code_kra.txt', 'w')
								f.write(str(code_kra))
								f.close()
							elif etap == 'kod1.8':
								code_118 = event.text
								f = open('kvartbot/code_118.txt', 'w')
								f.write(str(code_118))
								f.close()
							write_msg_no(event.user_id, 'Сохраненно')
							ws['B'+str(nomer)].value = '1'
						else:
							write_msg(event.user_id, 'Укажите 4 цифры от 0 до 9 (Пример: 1234)')

					#--------------------------------------------------------------------------------------------------------------------------
					# указывание прибыли
					elif etap == '1.1.1' or etap == '1.2.1'  or etap == '1.3.1' or etap == '1.4.1' or etap == '1.5.1' or etap == '1.6.1' or etap == '1.7.1' or etap == '1.8.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								if etap == '1.1.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.1.1'
								
							try:		
								if etap == '1.2.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.2.1'

							try:		
								if etap == '1.3.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.3.1'

							try:		
								if etap == '1.4.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.4.1'

							try:		
								if etap == '1.5.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.5.1'

							try:		
								if etap == '1.6.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.6.1'

							try:		
								if etap == '1.7.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.7.1'

							try:		
								if etap == '1.8.1':
									money(ws, etap, event.text, event.user_id)

							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.8.1'

							if ws['B'+str(nomer)].value == '1.1.1.1' or ws['B'+str(nomer)].value =='1.2.1.1' or ws['B'+str(nomer)].value =='1.3.1.1' or ws['B'+str(nomer)].value =='1.4.1.1' or ws['B'+str(nomer)].value =='1.5.1.1' or ws['B'+str(nomer)].value =='1.6.1.1' or ws['B'+str(nomer)].value =='1.7.1.1' or ws['B'+str(nomer)].value =='1.8.1.1':
								write_msg(event.user_id, 'Укажи кол-во дней')
					#--------------------------------------------------------------------------------------------------------------------------
					#указывние кол-ва дней
					elif etap == '1.1.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 1, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.1'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.1.1.1'
					#указывние кол-ва дней
					elif etap == '1.2.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 2, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.2'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.2.1.1'

					#указывние кол-ва дней
					elif etap == '1.3.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 3, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.3'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.3.1.1'

					#указывние кол-ва дней
					elif etap == '1.4.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 4, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.4'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.4.1.1'

					#указывние кол-ва дней
					elif etap == '1.5.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 5, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.5'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.5.1.1'

					#указывние кол-ва дней
					elif etap == '1.6.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 6, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.6'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.6.1.1'

					#указывние кол-ва дней
					elif etap == '1.7.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 7, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.7'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.7.1.1'

					#указывние кол-ва дней
					elif etap == '1.8.1.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								plus(ws, etap, 8, event.text, event.user_id)
								ws['B'+str(nomer)].value = '1.8'
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.8.1.1'
					#--------------------------------------------------------------------------------------------------------------------------
					# указывание разходов
					elif etap == '1.1.2' or etap == '1.2.2' or etap == '1.3.2' or etap == '1.4.2' or etap == '1.5.2' or etap == '1.6.2' or etap == '1.7.2' or etap == '1.8.2':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:

							try:
								if etap == '1.1.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.1.2'

							try:
								if etap == '1.2.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.2.2'

							try:
								if etap == '1.3.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.3.2'

							try:
								if etap == '1.4.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.4.2'

							try:
								if etap == '1.5.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.5.2'

							try:
								if etap == '1.6.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.6.2'

							try:
								if etap == '1.7.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.7.2'

							try:
								if etap == '1.8.2':
									money(ws, etap, event.text, event.user_id)
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '1.8.2'

							if ws['B'+str(nomer)].value == '1.1.2.1' or ws['B'+str(nomer)].value =='1.2.2.1'  or ws['B'+str(nomer)].value =='1.3.2.1' or ws['B'+str(nomer)].value =='1.4.2.1' or ws['B'+str(nomer)].value =='1.5.2.1' or ws['B'+str(nomer)].value =='1.6.2.1' or ws['B'+str(nomer)].value =='1.7.2.1'  or ws['B'+str(nomer)].value =='1.8.2.1':
								write_msg_naim(event.user_id, 'Укажи наименование')
					# указывание нименования
					#--------------------------------------------------------------------------------------------------------------------------
					elif etap == '1.1.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 1, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.1'

					elif etap == '1.2.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 2, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.2'

					elif etap == '1.3.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 3, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.3'

					elif etap == '1.4.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 4, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.4'

					elif etap == '1.5.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 5, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.5'

					elif etap == '1.6.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 6, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.6'

					elif etap == '1.7.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 7, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.7'

					elif etap == '1.8.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus(ws, etap, 8, event.text, event.user_id)
							ws['B'+str(nomer)].value = '1.8'
					#--------------------------------------------------------------------------------------------------------------------------
					elif etap == '0.2':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							try:
								money(ws, etap, event.text, event.user_id)
								write_msg(event.user_id, 'Укажи наименование')
							except Exception as e:
								write_msg(event.user_id, 'Укажи число')
								ws['B'+str(nomer)].value = '0.2'

					elif etap == '0.2.1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						else:
							minus_all(event)
							etap_NO(event.user_id, nomer, ws)
					#--------------------------------------------------------------------------------------------------------------------------
					#Выбираем квартиру
					elif etap == 'choice_edit':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						elif event.text == 'Моисеева' or event.text == 'Переверткина' or event.text == 'Ворошилова' or event.text == 'Короленко' or event.text == '20 лет' or event.text == 'Шишкова': 
							ws['B'+str(nomer)].value = 'choice_edit_1'
							ws['C'+str(nomer)].value = str(event.text)
							write_msg_day(event.user_id, 'Когда?')
						else:
							write_msg_object(event.user_id, 'Я вас не понял. Выберите объект')
					#Выбираем день
					elif etap == 'choice_edit_1':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						if event.text == 'Сегодня' or event.text == 'Вчера' or event.text == 'Позавчера':
							write_msg_pm_onli(event.user_id, 'Выберите прибыль или затраты')
							ws['B'+str(nomer)].value = 'choice_edit_2'
							ws['D'+str(nomer)].value = str(event.text)
						else:
							write_msg_day(event.user_id, 'Я вас не понял. Какой день редактировать?')
					#Выбираем +-
					elif etap == 'choice_edit_2':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						elif event.text == '+' or event.text == '-':
							rezult, kol_vo = edit(str(ws['C'+str(nomer)].value), str(ws['D'+str(nomer)].value), str(event.text))
							print(kol_vo)
							if kol_vo == 0:
								write_msg_no(event.user_id, rezult)
								etap_NO(event.user_id, nomer, ws)
							else:
								write_msg_universal(event.user_id, rezult, kol_vo)
								ws['B'+str(nomer)].value = 'choice_edit_3'
								ws['E'+str(nomer)].value = str(event.user_id)
						else:
							write_msg_pm_onli(event.user_id, 'Я вас не понял. Выберите прибыль или затраты')
					#Выбираем запись
					elif etap == 'choice_edit_3':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						elif exam(event.text) == True:
							ws['B'+str(nomer)].value = 'choice_edit_4'
							ws['F'+str(nomer)].value = str(event.text)
							write_msg_redact(event.user_id, 'Выберите что вы хотите сделать')
						else:
							write_msg_universal(event.user_id, 'Я тебя не понял. Выберите цифрой что редактировать', 12)

					elif etap == 'choice_edit_4':
						if event.text == 'Вернуться':
							etap_NO(event.user_id, nomer, ws)
						elif event.text == 'Удалить':
							# Удалить запись(название файла, день, прибыль или затраты, номер записи)
							rezult = delite_zapis(str(ws['C'+str(nomer)].value), str(ws['D'+str(nomer)].value), str(ws['E'+str(nomer)].value), str(ws['F'+str(nomer)].value))
							if rezult == 'NO':
								write_msg_no(event.user_id, 'Такой записи нет')
							else:	
								write_msg_no(event.user_id, 'Запись "' + str(rezult) + '" была удалена.')
							etap_NO(event.user_id, nomer, ws)
						elif event.text == 'Редактировать':
							write_msg(event.user_id, 'В разработке..')
						else:
							write_msg_redact(event.user_id, 'Я тебя не понял. Выберите что вы хотите сделать')
					#--------------------------------------------------------------------------------------------------------------------------
					elif etap == 'NOACCESS':
						write_msg_access(event.user_id, 'У вас нет доступа. Введите код.')
						ws['B'+str(nomer)].value = 'ACCESS'
					elif etap == 'ACCESS':
						if event.text == '2965':
							etap_NO(event.user_id, nomer, ws)
							f = open('kvartbot/admin_id.txt', 'a')
							f.write(str(event.user_id) + '\n')
							f.close()
						else:
							write_msg_access(event.user_id, 'Код неверный.')
					else:
						etap_NO(event.user_id, nomer, ws)
						
					wb.save('kvartbot/acc.xlsx')
					wb.close()

	except Exception as e:
		f = open('ErrorLog.txt', 'a')
		f.write(str(e))
		moscow = datetime.timezone(datetime.timedelta(hours=3))
		f.write('\n'+ str(datetime.datetime.now(moscow))+ '\n')
		f.close()	
		print(traceback.format_exc())		





